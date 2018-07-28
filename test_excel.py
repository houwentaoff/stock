import openpyxl
from openpyxl.styles import Font, Color
from openpyxl.styles import colors, PatternFill
# 新建
excel = openpyxl.Workbook()
# 创建表页

for i in range(1, 20):
    #for name in ['1', 'fsa', 'vv', 'xxx']:
    name = "股票表" + str(i)
    excel.create_sheet(name)
# 读取表页
print(excel.sheetnames)
# 获取当前激活页
ws = excel.active
# 设置表情页名
ws.title = "修改后的名字"
print(excel.sheetnames)
# 获取标签页
ws = excel['股票表2']
ws.title = 'asf'
print(excel.sheetnames)

# 遍历excel 页
for sheet in excel:
    print(sheet.title)
#不常用
# 增加8列
lcell = []
for i in range(1,8):
    cell = openpyxl.worksheet.write_only.WriteOnlyCell(ws, value=str(i))
    lcell.append(cell)

ws.append(lcell)
#常用
ws = excel.create_sheet(title="股票列表最后一页")
i = 0

cols = ["名字", "代码", "当前价格", "总市值"]
for col in cols:
    i += 1
    ws.cell(column=i, row=1, value=col)

#for col in range(1, 7):
#    ws.cell(column=col, row=1, value=""+str(col)+"列")
# 遍历1
for c in ws.columns:
    print(c, type(c))

# 遍历2
colmax = ws.max_column
print("最大列:", colmax)
for i in range(1, colmax+1):
    print(ws.cell(row=1, column=i).value)

# 增加row
rows = range(2, 100)
for r in range(2, 100):
    for i in range(1, colmax+1):
        ws.cell(r, i, value = "row:"+str(rows[r-2]) + ",col:"+cols[i-1])

excel.save('sss.xlsx')

wbr = openpyxl.load_workbook(filename='sss.xlsx')#, read_only=True)
ws = wbr['股票列表最后一页']
for row in ws.rows:
    for cell in row:
        print(cell.value)

# 查找名字为xxx的股票对于的代码
stockName = "row:12,col:名字"
for row in ws.rows:
    cell = row[cols.index("名字")]
    if cell.value == stockName:
        codecell = row[cols.index('代码')]
        codecell.font = Font(color=colors.RED)#Font(color='FFFF0000')
        print("找到: 代码为:", codecell.value)
        moneycell = row[cols.index('当前价格')]
        # 填充背景色
        moneycell.fill = PatternFill(start_color ='FFFF00', end_color = 'FFFF00', fill_type = 'solid')
        # 下面语句是无效的
        #moneycell.fill = PatternFill(bgColor=colors.RED)#"FFFF0000") 


wbr.save("sss.xlsx")
