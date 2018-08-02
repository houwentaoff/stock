# -*- coding:utf-8 -*- 
'''
Created on 2017/10/14
@author: Joy Hou
'''
import tushare as ts
import time
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import traceback

from functools import wraps
import sys

import openpyxl
from openpyxl.styles import Font, Color
from openpyxl.styles import colors, PatternFill

import progressbar

def singleton(cls):
    instances = {}

    @wraps(cls)
    def getinstance(*args, **kwargs):
        if cls not in instances:
            instances[cls] = cls(*args, **kwargs)
            return instances
        return instances

#@singleton
class Stocks(object):
    instance = None

    def __init__(self):
        self.allStocks = None
        self.codes = []
# 用print(Stock.GetInstance) 查看是否绑定
    @classmethod
    def GetInstance(cls):
        if cls.instance == None:
            cls.instance = Stocks()
        return cls.instance
    # 外部函数 和 该类无关系 不进行绑定
    @staticmethod
    def GetInstance1():
        if Stocks.instance == None:
            Stocks.instance = Stocks()
        return Stocks.instance

    def __check(self):
        for i in self.__stocks.index:
            nt = self.__stocks.loc[i].values[self.__getIdx('name')]
            for c in nt:
                if c == ' ':
                    #print('包含有空格:', nt)
                    nt = nt.replace(' ','')
                    self.__stocks.loc[i, 'name'] = nt# .values[self.__getIdx('name')] = nt
                    #nt = self.__stocks.loc[i].values[self.__getIdx('name')] 
                    #print('包含有空格:', nt)
                    break

    def GetStocks(self):
        if type(self.allStocks) == type(None):
            self. allStocks = ts.get_stock_basics()
            self.__check()

        return self.allStocks

    @property
    def __stocks(self):
        return self.GetStocks()

    def __getCols(self):
        return self.__stocks.columns.values.tolist()

    def __getIdx(self, name):
        cols = self.__getCols()
        return cols.index(name)

## public
    def GetCode(self, name):
        lcode = [] 
        for i in self.__stocks.index:
            nt = self.__stocks.loc[i].values[self.__getIdx('name')]
            if  name == nt:
                c = i #self.__stocks.loc[i].values[self.__getIdx('code')]
                lcode.append(c)
        return lcode

    def GetCodes(self):
        if len(self.codes) == 0:
            for i in self.__stocks.index:
                self.codes.append(i)
        return self.codes

    def GetName(self, code):
        return self.__stocks.loc[code, 'name']

    def GetStock(self, code):
        name = self.__stocks.loc[code, 'name']
        industry  = self.__stocks.loc[code, 'industry']
        pe = self.__stocks.loc[code, 'pe']
        outstanding = self.__stocks.loc[code, 'outstanding']
        total = self.__stocks.loc[code, 'totals']
        per = float(outstanding)/total
        s = StockSimple(code, name, industry, pe, outstanding, total, per)
        return s

    def IsST(self, code):
        name = self.__stocks.loc[code, 'name']
        if -1 != name.find('ST'):
            return True
        return False

    def GetPE(self, code):
        pe = self.__stocks.loc[code, 'pe']
        return pe

    def PrintInfo(self, code):
        name = self.__stocks.loc[code, 'name']
        industry  = self.__stocks.loc[code, 'industry']
        pe = self.__stocks.loc[code, 'pe']
        outstanding = self.__stocks.loc[code, 'outstanding']
        total = self.__stocks.loc[code, 'totals']
        per = float(outstanding)/total
        print (name, " ", industry,"市盈:", pe, " 流通:", outstanding, " 总股本:", total, " 流通占比:", per)

class Util(object):
    def WriteToxlsx(self, fileName="", objs=[]):
        wb = openpyxl.Workbook()
        wb.create_sheet("xx")
        ws = wb.active
        cols = ['代码', '名字', '行业', '市盈', '流通', '总股本', '流通占比']
        i = 0
        for col in cols:
            i += 1
            ws.cell(column = i, row=1, value=col)
        r = 2
        for obj in objs:
            for col in cols:
                ws.cell(r, cols.index(col) + 1, value = obj.GetValue(col))
            r += 1
        wb.save(fileName)

class StockSimple(object):
    def __init__(self, code = "", name = "", industry="", pe="", outstanding="", total="", per=0.0):
        self.code = code
        self.name = name
        self.industry = industry
        self.pe = pe
        self.outstanding = outstanding
        self.total = total
        self.per = per

    def GetValue(self, alias):
        if alias == '代码':
            return self.code
        if alias == '名字':
            return self.name
        if alias == '行业':
            return self.industry
        if alias == '市盈':
            return self.pe
        if alias == '流通':
            return self.outstanding
        if alias == '总股本':
            return self.total
        if alias == '流通占比':
            return self.per


    @property
    def Code(self):
        return self.code
    @property
    def CirRate(self):
        return str(self.per*100)+'%%'

class Algorithm(object):
    pbar = None
    
    def IsTRadeDay(self, day):
        validdates = ts.trade_cal()
        for d in validdates.values:
            if d[0] == day:
                return True
        return False
    
    def __get_day(self, endDay, interDay):
        validdates = ts.trade_cal()
        newDayIdx = 0
        oldday = ''
        
        for i in range(len(validdates)):
            if validdates.values[i][0] == endDay:
                newDayIdx = i
                break
        
        for i in range(newDayIdx, 0, -1):
            if validdates.values[i][1] == 1:
                interDay -= 1
                if interDay == 0:
                    oldday = validdates.values[i][0]
                    break
        return oldday


    def __get_oldday(self, now, interval):
        return self.__get_day(endDay = now, interDay = interval)
    '''
        validdates = ts.trade_cal()

        now = time.strftime('%Y-%m-%d',time.localtime(time1))

        inter_time = interval*24*3600
        old_time = time.strftime('%Y-%m-%d',time.localtime(time1-inter_time))
        return old_time
    '''
    def GetNDayPrice(self, code, day, n=30):

        old_day = self.__get_oldday(day, n)
        data = ts.get_hist_data(code, start=old_day, end=day, ktype='D')
        v_price = 0.0
        try:
            if type(data) == type(None) or type(data['close']) == type(None):
                return v_price
            count = len(data['close'])
            priceTotal = 0.0
            if count > n/2:
                for i in range(0, count):
                    priceTotal += data['close'][i]
                v_price = priceTotal/count
        
        except Exception as e:
            msg = traceback.format_exc()
            print(msg)
        return v_price
    # 长时间横盘,未上涨,
    ##
    # @brief 
    #
    # @param weight           横盘时间
    # @param amplitudeDot     中间振幅点数
    # @param maxjuli_30dPrice 距离30日均线 最大涨幅点数
    #
    # @return 
    def HengNoUp(self, weight=30, amplitudeDot=5, maxjuli_30dPrice=4, enableST=False, pemax=200.0):
        upDotRate = amplitudeDot/100.0
        curUpDotRate = 5/100.0
        perMaxJuli = maxjuli_30dPrice/100.0
        stocks = Stocks.GetInstance()
        codes = stocks.GetCodes()
        today=time.strftime('%Y-%m-%d',time.localtime(time.time()))
        old_day = self.__get_oldday(today, weight)
        print(old_day,'-', today, '间隔', weight, '交易日')
        li = []
        if Algorithm.pbar == None:
            widgets = ['',progressbar.Percentage(), ' ', progressbar.Bar('#'),' ', progressbar.Timer(),  
                               ' ', progressbar.ETA(), ' '] # FileTransferSpeed()]  
            Algorithm.pbar = progressbar.ProgressBar(widgets=widgets, maxval=len(codes))
            Algorithm.pbar.start()

        for v in codes:
            Algorithm.pbar.update(codes.index(v))
            if enableST == False and stocks.IsST(v) == True:
                continue
            if stocks.GetPE(v) > pemax:
                continue

            data = ts.get_hist_data(v, start=old_day, end=today, ktype='D')
            #print(data)
            try:
                
                if type(data) == type(None) or type(data['close']) == type(None):
                    continue
                count = len(data['close'])
                if count > weight/2:
                    begin = data['close'][count - 1]
                    valid_count = 0
                    avalid_count = 0
                    #for vv in data['close']:
                    for n in range(count-1, -1, -1):
                        prePrice = data['close'][n]
                        #print v, code_name_dic[v], vv, "n:", n, "count:", count, "valid:", valid_count
                        if n != 0:
                            #d = self.__get_oldday(today, n)
                            #v30 = self.GetNDayPrice(code=v, day=d, n=30)
                            #if v30 <= prePrice:
                            #    break
                            if abs(prePrice-begin) < upDotRate*begin:
                                valid_count += 1
                            else:
                                if avalid_count > 3:
                                    break
                                avalid_count += 1
                        else:
                            cur = data['close'][0] 
                            sec = data['close'][1] 
                            if cur > sec:
                                if cur - sec < curUpDotRate*sec:
                                    # 月线 > 当前价格n个点以内
                                    v30 = self.GetNDayPrice(code=v, day=today, n=30)
                                    if v30 > cur * (1+0.01) and v30 < cur * (1+perMaxJuli):
                                        valid_count += 1
                    #print "有效个数:", valid_count, " 无效个数:", avalid_count," 总数:", count                
                    if valid_count >= count-avalid_count:
                        name = stocks.GetName(v)
                        print (v, name, '开始价格', begin, '持续天数:', valid_count)
                        stocks.PrintInfo(v)
                        
                        tup = (v, data)
                        li.append(tup)

            except Exception as e:
                msg = traceback.format_exc()
                print(msg)
                #print ('异常', e)
        Algorithm.pbar.finish()
        return li     
    # 长时间横盘 突然上涨的股票
#
#获取当前日期算起Ｎ日的收盘价
# 
    ##
    # @brief 
    #
    # @param weight    横盘天数
    # @param updot      
    # @param curupdot  最近一天上涨点数
    # @param enableST  使能ST
    # @param pe        PE最大值
    #
    # @return          满足条件的股票
    def HengUp(self, weight=30, updot=4, curupdot=2, enableST=True, pemax=200.0):
        upDotRate = updot/100.0
        curUpDotRate = curupdot/100.0
        stocks = Stocks.GetInstance()
        codes = stocks.GetCodes()
        today=time.strftime('%Y-%m-%d',time.localtime(time.time()))
        old_day = self.__get_oldday(today, weight)
        print(old_day,'-', today, '间隔', weight, '交易日')
        li = []
        if Algorithm.pbar == None:
            widgets = ['',progressbar.Percentage(), ' ', progressbar.Bar('#'),' ', progressbar.Timer(),  
                               ' ', progressbar.ETA(), ' '] # FileTransferSpeed()]  
            Algorithm.pbar = progressbar.ProgressBar(widgets=widgets, maxval=len(codes))
            Algorithm.pbar.start()

        for v in codes:
            Algorithm.pbar.update(codes.index(v))
            if enableST == False and stocks.IsST(v) == True:
                continue
            if stocks.GetPE(v) > pemax:
                continue

            data = ts.get_hist_data(v, start=old_day, end=today, ktype='D')
            #print(data)
            try:
                
                if type(data) == type(None) or type(data['close']) == type(None):
                    continue
                count = len(data['close'])
                if count > weight/2:
                    begin = data['close'][count - 1]
                    valid_count = 0
                    avalid_count = 0
                    #for vv in data['close']:
                    for n in range(count-1, -1, -1):
                        prePrice = data['close'][n]
                        #print v, code_name_dic[v], vv, "n:", n, "count:", count, "valid:", valid_count
                        if n != 0:
                            if abs(prePrice-begin) < upDotRate*begin:
                                valid_count += 1
                            else:
                                if avalid_count > 3:
                                    break
                                avalid_count += 1
                        else:
                            cur = data['close'][0] 
                            sec = data['close'][1] 
                            if cur > sec:
                                if cur - sec > curUpDotRate*sec:
                                    valid_count += 1
                    #print "有效个数:", valid_count, " 无效个数:", avalid_count," 总数:", count                
                    if valid_count >= count-avalid_count:
                        name = stocks.GetName(v)
                        print (v, name, '开始价格', begin, '持续天数:', valid_count)
                        stocks.PrintInfo(v)
                        
                        tup = (v, data)
                        li.append(tup)

            except Exception as e:
                msg = traceback.format_exc()
                print(msg)
                #print ('异常', e)
        Algorithm.pbar.finish()
        return li        

alg = Algorithm()
#li = alg.HengUp(weight=30, updot=4, curupdot=2, enableST=False, pemax=200.0)
li = alg.HengNoUp(weight=20, amplitudeDot=5, maxjuli_30dPrice=4, enableST=False, pemax=200.0)
# (code, data)

stocks = Stocks.GetInstance()
ls = []
for tup in li:
    code = tup[0]
    sobj = stocks.GetStock(code)
    ls.append(sobj)
    #stocks.PrintInfo(code)
    #tup[1]['close'].plot()
    #plt.show()

ls.sort(key=lambda obj: obj.pe, reverse=False)    
print("sort:")
print("write to xlsx")
util = Util() 
util.WriteToxlsx(fileName="xxx.xlsx", objs=ls)
for s in ls:
    stocks.PrintInfo(s.Code)
